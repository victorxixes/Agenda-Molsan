from typing import Dict, List
from openpyxl import load_workbook
from fastapi import UploadFile
from sqlalchemy.orm import Session

from backend.app.ctn.models import Notaria

# Mapeo entre cabeceras del Excel y campos del modelo
HEADER_TO_FIELD: Dict[str, str] = {
    "Código": "codigo",
    "Nombre": "nombre",
    "Apellidos": "apellidos",
    "NIF": "nif",
    "Teléfono": "telefono",
    "Departamento cancelaciones": "departamento_cancelaciones",
    "Departamento copias": "departamento_copias",
    "Otros departamentos": "otros_departamentos",
    "CP": "cp",
    "Provincia": "provincia",
    "Municipio": "municipio",
    "VC": "vc",
    "Apoderado": "apoderado",
    "Apoderado S": "apoderado_s",
    "Observación": "observacion",
}

def importar_ctn_desde_excel(db: Session, file: UploadFile) -> int:
    """
    Lee el Excel de CTN y rellena la tabla ctn_notarios
    tal y como está el fichero, sin transformar nada.
    - Ignora la primera fila (título)
    - Usa la segunda fila como cabeceras
    - Inserta desde la tercera fila en adelante
    """

    contents = file.file.read()
    wb = load_workbook(filename=None, data=contents)
    ws = wb.active

    filas = list(ws.iter_rows(values_only=True))
    if len(filas) < 3:
        return 0

    # Fila 1: título (ignorar)
    # Fila 2: cabeceras
    headers = filas[1]
    data_rows = filas[2:]

    # Opcional: limpiar tabla antes de importar
    db.query(Notaria).delete()

    insertados = 0

    for row in data_rows:
        if all(cell is None for cell in row):
            continue

        notaria_data: Dict[str, str] = {}

        for idx, header in enumerate(headers):
            if header is None:
                continue
            if header not in HEADER_TO_FIELD:
                continue

            field_name = HEADER_TO_FIELD[header]
            valor = row[idx] if idx < len(row) else None

            # Tal cual: todo a string si viene algo, None si está vacío
            notaria_data[field_name] = None if valor is None else str(valor)

        notaria = Notaria(**notaria_data)
        db.add(notaria)
        insertados += 1

    db.commit()
    return insertados
